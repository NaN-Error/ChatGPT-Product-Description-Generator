import os
import shutil
import time
import openpyxl
from openai import OpenAI
import openai
from dotenv import load_dotenv
load_dotenv()

data_modified = False  # Global flag

# Initialize the OpenAI client
client = OpenAI(api_key=os.environ.get('OPENAI_API_KEY'))
print("API client initialized.")

def rewrite_description(raw_description):
    """
    Rewrites a product description using OpenAI's API.

    Args:
    - raw_description (str): The original product description text.

    Returns:
    - str: The rewritten product description.

    Raises:
    - Exception: If all API call attempts fail.
    """
    global data_modified

    # Check if the raw description is None (empty or not provided)
    if raw_description is None:
        print("Error: Raw description is None")
        return None

    print("Rewriting description for raw data:", raw_description[:50])  # Print first 50 characters of raw description

    # Detailed instructions for rewriting the description
    instructions = """
        Rewrite the product description following these guidelines:
        - Translate to Spanish
        - Give a short but descriptive title for the product
        - Give me the product specifications in a concise format
        - Give me a brief summary of key product features
        - Give me a description focused on how the product can help or be useful to the customer
        - No bold text in any part of the content or title
        - Double space between each point
        - Example of guidelines:
            Lámpara de Escritorio LED de Doble Cabeza con Pinza y Cuello Flexible

            Especificaciones:
            - Material: Aluminio.
            - Rango de Ajuste: Barra de luz ajustable dentro de 180°.
            - Modos de Color: 5 rangos (2700k a 6500K).
            - Niveles de Brillo: 5 niveles (10% a 100%).
            - Cuidado de los Ojos: CRI>80, 1400lm.
            - Cuello de Cisne: Flexible, 360° ajustable.
            - Altura Máxima de Trabajo: 25 pulgadas.
            - Potencia: 12W.
            - Cable: Adaptador ETL de 78 pulgadas.

            Características:
            - Diseño de Doble Cabeza: Amplía el ángulo de iluminación, ideal para múltiples usos.
            - Ajustes Personalizables: Variedad en modos de color y niveles de brillo para adaptarse a diferentes necesidades.
            - Cuidado de los Ojos: Luz suave y cómoda, sin deslumbramiento ni parpadeo.
            - Función de Memoria: Recuerda el último ajuste de brillo y modo utilizado.
            - Cuello de Cisne Flexible: Ajuste fácil para evitar la luz directa y minimizar la fatiga visual.
            - Construcción Duradera: Hecha de aluminio resistente, con pinza firme y antideslizante.

            Descripción:
            La Lámpara de Escritorio LED de Doble Cabeza con Pinza y Cuello Flexible es una solución de iluminación versátil y eficiente para tu espacio de trabajo. Su diseño de doble cabeza permite extender el ángulo de iluminación, ofreciendo una cobertura amplia para diferentes actividades como leer, estudiar o dibujar. Fabricada en aluminio, es ligera y resistente, con una pinza equipada con almohadillas protectoras para un agarre seguro sin dañar la superficie de la mesa.

            Con cinco modos de color y cinco niveles de brillo, puedes personalizar la iluminación según tus preferencias y necesidades. La lámpara está diseñada para cuidar tus ojos, reduciendo el deslumbramiento y el parpadeo. El cuello de cisne flexible permite ajustar fácilmente la posición de la luz, mientras que la función de memoria es práctica para recordar tus ajustes preferidos. Ideal para una variedad de entornos, esta lámpara combina funcionalidad y estilo para mejorar tu experiencia de iluminación.
        """

    # Combine instructions with the raw product description
    prompt = instructions + "Text to rewrite:\n" + raw_description
    print("Prompt for API call prepared.")

    for attempt in range(3):  # Number of retries
        try:
            print(f"API call attempt {attempt + 1}")
            completion = client.chat.completions.create(
                model="gpt-3.5-turbo",  # Using GPT-4
                messages=[{"role": "user", "content": prompt}]
            )
            if completion.choices:
                choice = completion.choices[0]
                rewritten_content = choice.message.content.strip()
                data_modified = True
                return rewritten_content
        except Exception as e:
            print(f"An error occurred: {e}")
            if attempt < 2:
                print("Retrying after 10 seconds...")
                time.sleep(10)
            else:
                print("All retries failed.")
                raise


def process_excel_file(file_path):
    global data_modified

    backup_path = file_path.replace('.xlsx', '_backup.xlsx')

    # Check if the file exists
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return

    # Create a backup of the original file
    shutil.copy(file_path, backup_path)
    print(f"Backup created at: {backup_path}")

    try:
        # Function to reopen the workbook
        def reopen_workbook():
            nonlocal workbook, sheet
            workbook.save(file_path)
            workbook.close()
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

        # Open the workbook
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Identify the column numbers for required columns
        product_desc_col = None
        raw_desc_col = None
        sold_col = None
        damaged_col = None
        personal_col = None
        cancelled_col = None

        header_row = next(sheet.iter_rows(min_row=1, max_row=1))
        for cell in header_row:
            if cell.value == 'Product Description':
                product_desc_col = cell.column
            elif cell.value == 'Raw Product Description':
                raw_desc_col = cell.column
            elif cell.value == 'Sold':
                sold_col = cell.column
            elif cell.value == 'Damaged':
                damaged_col = cell.column
            elif cell.value == 'Personal':
                personal_col = cell.column
            elif cell.value == 'Cancelled Order':
                cancelled_col = cell.column

        if not all([product_desc_col, raw_desc_col, sold_col, damaged_col, personal_col, cancelled_col]):
            raise ValueError("Required columns not found in the Excel sheet")

        save_interval = 5  # Save after every 5 rows processed
        for row in range(2, sheet.max_row + 1):
            product_description = sheet.cell(row=row, column=product_desc_col).value
            sold_status = sheet.cell(row=row, column=sold_col).value
            damaged_status = sheet.cell(row=row, column=damaged_col).value
            personal_status = sheet.cell(row=row, column=personal_col).value
            cancelled_status = sheet.cell(row=row, column=cancelled_col).value

            should_rewrite = (
                (not product_description or product_description == "No Product Description At The Moment.") and
                sold_status != "YES" and damaged_status != "YES" and
                personal_status != "YES" and cancelled_status != "YES"
            )

            if should_rewrite:
                raw_description = sheet.cell(row=row, column=raw_desc_col).value
                if raw_description:
                    rewritten_description = rewrite_description(raw_description)
                    if rewritten_description:
                        sheet.cell(row=row, column=product_desc_col).value = rewritten_description
                        print(f"Row {row} updated with new description.")
                    else:
                        print(f"Skipping row {row} due to empty rewritten description.")
                else:
                    print(f"Skipping row {row} due to empty Raw Product Description.")
            else:
                print(f"Skipping row {row} as it does not meet criteria for update.")

            # Save the workbook every 'save_interval' rows
            if row % save_interval == 0 and data_modified:
                reopen_workbook()  # Save and reopen the workbook
                print(f"Saved progress at row {row}")
                data_modified = False  # Reset the flag after saving

        # Final save and replace original file if the process completes successfully
        workbook.save(file_path)
        workbook.close()
        print("Updates complete. Original file replaced with updated data.")
    except KeyboardInterrupt:
        print("Keyboard Interrupt detected. Closing the file without saving.")
        try:
            workbook.close()
            print("File closed without saving.")
        except ValueError:
            print("Workbook was already closed. No further action needed.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Example usage
file_path = 'R:\\Inventory Management.xlsx'
process_excel_file(file_path)

